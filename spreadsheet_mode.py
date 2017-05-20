import openpyxl
import sys
import os
import time

class Bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def getAveragesInRange():
    try:
        start = raw_input("Enter start date (dd.mm.yyyy): ")
        end = raw_input("Enter end date (dd.mm.yyyy): ")
        worksheet = raw_input("Enter name of the sheet: ")
        #start of part for getting data from sheet
        wb = openpyxl.load_workbook("SOBm-data.xlsx")
        sheet = wb.get_sheet_by_name(worksheet)
        average_happiness = 0
        average_energy = 0
        average_focus = 0
        for char in xrange(67, 70):
            avr = 0.0
            for row in xrange(2, sheet.max_row):
                if  sheet[str(chr(char)) + str(row)].value == None:
                    break
                #print sheet[str(chr(char)) + str(row)].value
                if sheet['A' + str(row)].value >= start and sheet['A' + str(row)].value <= end:
                    avr += sheet[str(chr(char)) + str(row)].value
            if char == 67:
                average_energy = avr / (sheet.max_row - 1)
            elif char == 68: 
                average_happiness = avr / (sheet.max_row -1)
            else:
                average_focus = avr / (sheet.max_row -1)
        print "Average happiness: {:.2f}\nAverage energy: {:.2f}\nAverage focus: {:.2f}".format(average_energy, average_happiness, average_focus)

    except IOError:
        print Bcolors.WARNING + "No such file or directory: {}\nProgram exited!".format(filename) + Bcolors.ENDC
        sys.exit()
    except KeyError:
        print Bcolors.WARNING + "Worksheet '{}' does not exist.\nProgram exited!".format(worksheet) + Bcolors.ENDC
        sys.exit()
    except Exception, e:
        print Bcolors.FAIL + repr(e) + Bcolors.ENDC
        print Bcolors.WARNING + "Program exited!" + Bcolors.ENDC
        sys.exit()

def write(sheet, happiness, energy, focus, spreadsheet_line_to_write):
    try:
        position = "A" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. A2, from column name + spreadsheet_line_to_write (line number)
        sheet[position] = time.strftime("%d.%m.%Y") #write date (dd.mm.yyyy) to sheet on spreadsheet_line_to_write eg. A2

        position = "B" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. B2, from column name + spreadsheet_line_to_write (line number)
        sheet[position] = time.strftime("%H:%M:%S") #write time (hh:mm:ss) to sheet on spreadsheet_line_to_write eg. B2

        position = "C" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. C2, from column name + spreadsheet_line_to_write (line number)
        sheet[position] = energy #write energy value to sheet on spreadsheet_line_to_write eg. C2

        position = "D" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. D2, from column name + spreadsheet_line_to_write (line number)
        sheet[position] = happiness #write happiness value to sheet on spreadsheet_line_to_write eg. D2

        position = "E" + str(spreadsheet_line_to_write) #get spreadsheet_line_to_write eg. E2, from column name + spreadsheet_line_to_write (line number)
        sheet[position] = focus #write focus value to sheet on spreadsheet_line_to_write eg. E2

        spreadsheet_line_to_write += 1 #increase value of spreadsheet_line_to_write (sheet line) by one
        sheet['G2'] = spreadsheet_line_to_write #write spreadsheet_line_to_write value to the shell on spreadsheet_line_to_write G2

        print Bcolors.OKGREEN + "Data was successfuly entered into a sheet at {}!".format(time.strftime("%H:%M:%S")) + Bcolors.ENDC
        
    except Exception, e:
        print Bcolors.WARNING + "Failed enetering data to the spreadsheet!" + Bcolors.ENDC
        print Bcolors.FAIL + repr(e) + Bcolors.ENDC
        print Bcolors.WARNING + "Program exited!" + Bcolors.ENDC
        sys.exit()

def saveSpreadsheet(wb):
    try:
        wb.save("SOBm-data.xlsx") #save spreadsheet to /home/bostjan/...
        print Bcolors.OKGREEN + "File saved successfully!" + Bcolors.ENDC

    except Exception, e:
        print Bcolors.WARNING + "Failed saving the file!" + Bcolors.ENDC
        print Bcolors.FAIL + repr(e) + Bcolors.ENDC
        print Bcolors.WARNING + "Program exited!" + Bcolors.ENDC
        sys.exit()
        
def getAverages(worksheet):
 try:
     #start of part for getting data from sheet
     wb = openpyxl.load_workbook("SOBm-data.xlsx")
     sheet = wb.get_sheet_by_name(worksheet)
     average_happiness = 0
     average_energy = 0
     average_focus = 0
     for char in xrange(67, 70):
         avr = 0.0
         for row in xrange(2, sheet.max_row):
             if  sheet[str(chr(char)) + str(row)].value == None:
                 break
             #print sheet[str(chr(char)) + str(row)].value
             avr += sheet[str(chr(char)) + str(row)].value
         if char == 67:
             average_energy = avr / (sheet.max_row - 1)
         elif char == 68: 
             average_happiness = avr / (sheet.max_row -1)
         else:
             average_focus = avr / (sheet.max_row -1)
     print "Average happiness: {:.2f}\nAverage energy: {:.2f}\nAverage focus: {:.2f}".format(average_energy, average_happiness, average_focus)

 except IOError:
     print Bcolors.WARNING + "No such file or directory: {}\nProgram exited!".format(filename) + Bcolors.ENDC
     sys.exit()
 except KeyError:
     print Bcolors.WARNING + "Worksheet '{}' does not exist.\nProgram exited!".format(worksheet) + Bcolors.ENDC
     sys.exit()
 except Exception, e:
     print Bcolors.FAIL + repr(e) + Bcolors.ENDC
     print Bcolors.WARNING + "Program exited!" + Bcolors.ENDC
     sys.exit()

def createSpreadsheet():
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "data"
        sheet["A1"] = "Date"
        sheet["B1"] = "Time"
        sheet["C1"] = "Energy"
        sheet["D1"] = "Happiness"
        sheet["E1"] = "Focus"
        sheet["G1"] = "Spreadsheet_Line_To_Write"
        sheet["G2"] = 2
        wb.save("SOBm-data.xlsx")
        print Bcolors.OKGREEN + "Spreadsheet successfully crated!" + Bcolors.ENDC
        
    except Exception, e:
        print Bcolors.WARNING + "Failed crating a spreadsheet!" + Bcolors.ENDC
        print Bcolors.FAIL + repr(e) + Bcolors.ENDC
        print Bcolors.WARNING + "Program exited!" + Bcolors.ENDC
        sys.exit()

def main(values, sheetName):
    try:
        wb = openpyxl.load_workbook("SOBm-data.xlsx")
        sheet = wb.get_sheet_by_name(sheetName)
        spreadsheet_line_to_write = sheet['G2'].value
        write(sheet, values[0], values[1], values[2], spreadsheet_line_to_write)
        saveSpreadsheet(wb)

    except IOError:
        print Bcolors.WARNING + "No such file or directory: {}\nProgram exited!".format("SOB-data.xslx") + Bcolors.ENDC
        sys.exit()
    except KeyError:
        print Bcolors.WARNING + "Worksheet '{}' does not exist.\nProgram exited!".format(sheetName) + Bcolors.ENDC
        sys.exit()
    except Exception, e:
        print Bcolors.FAIL + repr(e) + Bcolors.ENDC
        print Bcolors.WARNING + "Program exited!" + Bcolors.ENDC
        sys.exit()
